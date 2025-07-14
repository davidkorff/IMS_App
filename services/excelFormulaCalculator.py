#!/usr/bin/env python3
"""
Excel Formula Calculator
This script uses openpyxl to open an Excel file, populate data, calculate formulas,
and save the result. This is a workaround for ExcelJS's lack of formula calculation.
"""

import sys
import base64
import json
import openpyxl
from openpyxl import load_workbook
import tempfile
import os

def calculate_excel_formulas(excel_base64, data_mappings):
    """
    Load Excel file, populate data, calculate formulas, and return updated file
    
    Args:
        excel_base64: Base64 encoded Excel file
        data_mappings: JSON string with data to populate
    
    Returns:
        Base64 encoded Excel file with calculated formulas
    """
    try:
        # Decode base64 to bytes
        excel_bytes = base64.b64decode(excel_base64)
        
        # Parse data mappings
        data = json.loads(data_mappings)
        
        # Create temporary file
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
            tmp_file.write(excel_bytes)
            tmp_path = tmp_file.name
        
        # Load workbook with data_only=False to preserve formulas
        wb = load_workbook(tmp_path, data_only=False)
        
        # Find submission_data sheet
        submission_sheet = None
        for sheet_name in ['submission_data', 'Submission_Data', 'SubmissionData', 'Data']:
            if sheet_name in wb.sheetnames:
                submission_sheet = wb[sheet_name]
                break
        
        if not submission_sheet:
            raise Exception("Could not find submission_data worksheet")
        
        print(f"Found submission sheet: {submission_sheet.title}", file=sys.stderr)
        
        # Populate data based on common patterns
        # Pattern 1: Field names in column A, values in column B
        for row in range(1, 100):  # Check first 100 rows
            cell_a = submission_sheet.cell(row=row, column=1)
            if cell_a.value and isinstance(cell_a.value, str):
                field_name = cell_a.value.strip()
                # Try multiple variations of the field name
                for key in data:
                    if (key.lower() == field_name.lower() or 
                        key.replace('_', ' ').lower() == field_name.lower() or
                        key.replace('_', '').lower() == field_name.replace(' ', '').lower()):
                        submission_sheet.cell(row=row, column=2).value = data[key]
                        print(f"Set {field_name} = {data[key]}", file=sys.stderr)
                        break
        
        # Pattern 2: Direct cell mappings (if specified)
        if 'cell_mappings' in data:
            for cell_ref, value in data['cell_mappings'].items():
                submission_sheet[cell_ref] = value
                print(f"Set cell {cell_ref} = {value}", file=sys.stderr)
        
        # Force formula recalculation by setting calculation mode
        wb.calculation.calcMode = 'automatic'
        
        # Important: We need to trigger recalculation by modifying cells
        # OpenPyXL doesn't actually calculate formulas, but we can try to 
        # make Excel calculate them on next open
        for sheet in wb.worksheets:
            sheet.sheet_properties.tabColor = None  # Trigger dirty flag
        
        # Save workbook with formulas intact
        output_path = tmp_path + '_output.xlsx'
        wb.save(output_path)
        
        # Try to use Excel COM if available (Windows only)
        calculated_path = None
        try:
            import win32com.client
            print("Using Excel COM for calculation...", file=sys.stderr)
            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            
            workbook = excel.Workbooks.Open(os.path.abspath(output_path))
            excel.Calculate()  # Force full calculation
            
            calculated_path = tmp_path + '_calculated.xlsx'
            workbook.SaveAs(os.path.abspath(calculated_path))
            workbook.Close()
            excel.Quit()
            
            # Use the COM-calculated file
            output_path = calculated_path
            print("Excel COM calculation successful", file=sys.stderr)
            
        except Exception as e:
            print(f"Excel COM not available: {e}", file=sys.stderr)
            print("Warning: Formulas may not be calculated", file=sys.stderr)
        
        # Now reload to get values (whether calculated or not)
        wb_calculated = load_workbook(output_path, data_only=True)
        
        # Log calculated values from key cells
        if 'IMS_TAGS' in wb_calculated.sheetnames:
            ims_sheet = wb_calculated['IMS_TAGS']
            for cell_ref in ['B6', 'B7', 'B8', 'B9', 'B10']:
                value = ims_sheet[cell_ref].value
                if value:
                    print(f"IMS_TAGS!{cell_ref} = {value}", file=sys.stderr)
        
        # Save final file with calculated values
        final_path = tmp_path + '_final.xlsx'
        wb_calculated.save(final_path)
        
        # Read and encode result
        with open(final_path, 'rb') as f:
            result_bytes = f.read()
        
        # Cleanup temp files
        for path in [tmp_path, output_path, final_path]:
            if os.path.exists(path):
                os.remove(path)
        
        if calculated_path and os.path.exists(calculated_path):
            os.remove(calculated_path)
        
        # Return base64 encoded result
        return base64.b64encode(result_bytes).decode('utf-8')
        
    except Exception as e:
        print(f"Error: {str(e)}", file=sys.stderr)
        raise

def calculate_excel_formulas_from_files(excel_path, data_path):
    """
    Load Excel file and data from file paths instead of base64
    """
    try:
        # Read Excel file
        with open(excel_path, 'rb') as f:
            excel_bytes = f.read()
        
        # Read data JSON
        with open(data_path, 'r') as f:
            data_json = f.read()
        
        # Convert to base64 and process
        excel_base64 = base64.b64encode(excel_bytes).decode('utf-8')
        return calculate_excel_formulas(excel_base64, data_json)
        
    except Exception as e:
        print(f"Error reading files: {str(e)}", file=sys.stderr)
        raise

if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python excelFormulaCalculator.py <excel_file_path> <data_json_path>", file=sys.stderr)
        sys.exit(1)
    
    try:
        result = calculate_excel_formulas_from_files(sys.argv[1], sys.argv[2])
        print(result)  # Output base64 result to stdout
    except Exception as e:
        print(f"ERROR: {str(e)}", file=sys.stderr)
        sys.exit(1)