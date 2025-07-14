#!/usr/bin/env python3
"""
Direct Excel Calculator using xlwings or win32com
This provides the most accurate formula calculation
"""

import sys
import os
import json
import base64

def calculate_excel_with_com(excel_path, data_path):
    """
    Use Excel COM automation (Windows only) for perfect calculation
    """
    try:
        import win32com.client
        print("Using Excel COM automation", file=sys.stderr)
        
        # Read data mapping
        with open(data_path, 'r') as f:
            data = json.loads(f.read())
        
        # Create Excel instance
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        
        try:
            # Open workbook
            workbook = excel.Workbooks.Open(os.path.abspath(excel_path))
            
            # Find submission_data sheet
            submission_sheet = None
            for sheet in workbook.Worksheets:
                if sheet.Name.lower() in ['submission_data', 'submission data', 'submissiondata', 'data']:
                    submission_sheet = sheet
                    break
            
            if submission_sheet:
                print(f"Found sheet: {submission_sheet.Name}", file=sys.stderr)
                
                # Populate cells from mapping
                if 'cell_mappings' in data:
                    for cell_ref, value in data['cell_mappings'].items():
                        submission_sheet.Range(cell_ref).Value = value
                        print(f"Set {cell_ref} = {value}", file=sys.stderr)
            
            # Force calculation
            excel.Calculate()
            excel.CalculateFull()
            
            # Save to new file
            output_path = excel_path + '_calculated.xlsx'
            workbook.SaveAs(os.path.abspath(output_path))
            
            # Check IMS_TAGS sheet for premium
            try:
                ims_tags_sheet = None
                for sheet in workbook.Worksheets:
                    if sheet.Name.upper() == 'IMS_TAGS':
                        ims_tags_sheet = sheet
                        break
                
                if ims_tags_sheet:
                    print(f"Found IMS_TAGS sheet", file=sys.stderr)
                    # Check B6 specifically for premium
                    b6_value = ims_tags_sheet.Range('B6').Value
                    if b6_value and isinstance(b6_value, (int, float)) and b6_value > 0:
                        print(f"Found premium in IMS_TAGS!B6: {b6_value}", file=sys.stderr)
            except Exception as e:
                print(f"Error checking IMS_TAGS: {e}", file=sys.stderr)
            
            # Close workbook
            workbook.Close(SaveChanges=False)
            
            # Read result
            with open(output_path, 'rb') as f:
                result_bytes = f.read()
            
            # Cleanup
            if os.path.exists(output_path):
                os.remove(output_path)
            
            return base64.b64encode(result_bytes).decode('utf-8')
            
        finally:
            excel.Quit()
            
    except ImportError:
        print("Excel COM not available (Windows with Excel required)", file=sys.stderr)
        print("Falling back to xlwings...", file=sys.stderr)
        return calculate_excel_with_xlwings(excel_path, data_path)

def calculate_excel_with_xlwings(excel_path, data_path):
    """
    Use xlwings for cross-platform Excel automation
    """
    try:
        import xlwings as xw
        print("Using xlwings for calculation", file=sys.stderr)
        
        # Read data mapping
        with open(data_path, 'r') as f:
            data = json.loads(f.read())
        
        # Open workbook
        app = xw.App(visible=False)
        try:
            wb = app.books.open(excel_path)
            
            # Find submission_data sheet
            submission_sheet = None
            for sheet in wb.sheets:
                if sheet.name.lower() in ['submission_data', 'submission data', 'submissiondata', 'data']:
                    submission_sheet = sheet
                    break
            
            if submission_sheet:
                print(f"Found sheet: {submission_sheet.name}", file=sys.stderr)
                
                # Populate cells
                if 'cell_mappings' in data:
                    for cell_ref, value in data['cell_mappings'].items():
                        submission_sheet.range(cell_ref).value = value
                        print(f"Set {cell_ref} = {value}", file=sys.stderr)
            
            # Force calculation
            app.calculate()
            
            # Save calculated workbook
            output_path = excel_path + '_calculated.xlsx'
            wb.save(output_path)
            wb.close()
            
            # Read result
            with open(output_path, 'rb') as f:
                result_bytes = f.read()
            
            # Cleanup
            if os.path.exists(output_path):
                os.remove(output_path)
            
            return base64.b64encode(result_bytes).decode('utf-8')
            
        finally:
            app.quit()
            
    except ImportError:
        print("xlwings not available. Install with: pip install xlwings", file=sys.stderr)
        print("Falling back to openpyxl (WARNING: NO FORMULA CALCULATION)", file=sys.stderr)
        return calculate_excel_fallback(excel_path, data_path)

def calculate_excel_fallback(excel_path, data_path):
    """
    Fallback using openpyxl (no formula calculation)
    """
    from openpyxl import load_workbook
    
    print("WARNING: Using openpyxl fallback - formulas won't calculate", file=sys.stderr)
    print("For formula calculation, install xlwings or use Windows with Excel", file=sys.stderr)
    
    # Read data mapping
    with open(data_path, 'r') as f:
        data = json.loads(f.read())
    
    # Load workbook
    wb = load_workbook(excel_path, data_only=False)
    
    # Find submission_data sheet
    submission_sheet = None
    for sheet_name in ['submission_data', 'Submission_Data', 'SubmissionData', 'Data']:
        if sheet_name in wb.sheetnames:
            submission_sheet = wb[sheet_name]
            break
    
    if submission_sheet:
        # Populate data
        if 'cell_mappings' in data:
            for cell_ref, value in data['cell_mappings'].items():
                submission_sheet[cell_ref] = value
                print(f"Set {cell_ref} = {value}", file=sys.stderr)
    
    # Save result
    output_path = excel_path + '_output.xlsx'
    wb.save(output_path)
    
    # Read and return as base64
    with open(output_path, 'rb') as f:
        result_bytes = f.read()
    
    # Cleanup
    if os.path.exists(output_path):
        os.remove(output_path)
    
    return base64.b64encode(result_bytes).decode('utf-8')

if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python pythonExcelCalculator.py <excel_file_path> <data_json_path>", file=sys.stderr)
        sys.exit(1)
    
    try:
        # Try COM first (most accurate), then xlwings, then fallback
        result = calculate_excel_with_com(sys.argv[1], sys.argv[2])
        print(result)  # Output base64 result to stdout
    except Exception as e:
        print(f"ERROR: {str(e)}", file=sys.stderr)
        sys.exit(1)