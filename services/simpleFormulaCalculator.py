#!/usr/bin/env python3
"""
Simple Excel Formula Calculator using formulas package
This provides better formula calculation than openpyxl
"""

import sys
import os
import json
import base64
from openpyxl import load_workbook

def calculate_with_formulas(excel_path, data_path):
    """
    Use formulas package for better calculation support
    """
    try:
        # First check if formulas package is available
        try:
            import formulas
            print("Using formulas package for calculation", file=sys.stderr)
        except ImportError:
            print("formulas package not installed. Install with: pip install formulas[all]", file=sys.stderr)
            # Fall back to openpyxl only
            return calculate_with_openpyxl_only(excel_path, data_path)
        
        # Read data mapping
        with open(data_path, 'r') as f:
            data = json.loads(f.read())
        
        # Load Excel file
        wb = load_workbook(excel_path, data_only=False)
        
        # Find submission_data sheet
        submission_sheet = None
        for sheet_name in ['submission_data', 'Submission_Data', 'SubmissionData', 'Data']:
            if sheet_name in wb.sheetnames:
                submission_sheet = wb[sheet_name]
                break
        
        if not submission_sheet:
            raise Exception("Could not find submission_data worksheet")
        
        # Populate data
        print(f"Populating data in {submission_sheet.title}", file=sys.stderr)
        
        # Direct cell mappings
        if 'cell_mappings' in data:
            for cell_ref, value in data['cell_mappings'].items():
                submission_sheet[cell_ref] = value
                print(f"Set {cell_ref} = {value}", file=sys.stderr)
        
        # Save populated workbook
        temp_populated = excel_path + '_populated.xlsx'
        wb.save(temp_populated)
        
        # Use formulas package to calculate
        print("Calculating formulas with formulas package...", file=sys.stderr)
        
        # Create Excel model
        excel_model = formulas.ExcelModel()
        excel_model.loads(temp_populated)
        excel_model.calculate()
        
        # Save calculated result
        temp_calculated = excel_path + '_calculated.xlsx'
        excel_model.write(dirpath=os.path.dirname(temp_calculated))
        
        # Read result and return as base64
        with open(temp_calculated, 'rb') as f:
            result_bytes = f.read()
        
        # Cleanup
        for path in [temp_populated, temp_calculated]:
            if os.path.exists(path):
                os.remove(path)
        
        return base64.b64encode(result_bytes).decode('utf-8')
        
    except Exception as e:
        print(f"Formulas calculation error: {str(e)}", file=sys.stderr)
        # Fall back to openpyxl only
        return calculate_with_openpyxl_only(excel_path, data_path)

def calculate_with_openpyxl_only(excel_path, data_path):
    """
    Fallback using only openpyxl (formulas won't calculate)
    """
    try:
        # Read data mapping
        with open(data_path, 'r') as f:
            data = json.loads(f.read())
        
        # Load workbook
        wb = load_workbook(excel_path, data_only=False)
        
        # Find submission_data sheet
        submission_sheet = None
        for sheet_name in ['submission_data', 'Submission_Data', 'SubmissionData', 'Data']:
            if sheet_name in wb.sheetnames:
                submission_sheet = wb[sheet_name]
                break
        
        if submission_sheet:
            # Populate data
            if 'cell_mappings' in data:
                for cell_ref, value in data['cell_mappings'].items():
                    submission_sheet[cell_ref] = value
                    print(f"Set {cell_ref} = {value}", file=sys.stderr)
        
        # Save result
        temp_output = excel_path + '_output.xlsx'
        wb.save(temp_output)
        
        print("WARNING: Formulas not calculated (install formulas package for calculation)", file=sys.stderr)
        
        # Read and return as base64
        with open(temp_output, 'rb') as f:
            result_bytes = f.read()
        
        # Cleanup
        if os.path.exists(temp_output):
            os.remove(temp_output)
        
        return base64.b64encode(result_bytes).decode('utf-8')
        
    except Exception as e:
        print(f"Error: {str(e)}", file=sys.stderr)
        raise

if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python simpleFormulaCalculator.py <excel_file_path> <data_json_path>", file=sys.stderr)
        sys.exit(1)
    
    try:
        result = calculate_with_formulas(sys.argv[1], sys.argv[2])
        print(result)  # Output base64 result to stdout
    except Exception as e:
        print(f"ERROR: {str(e)}", file=sys.stderr)
        sys.exit(1)